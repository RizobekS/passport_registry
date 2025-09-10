# apps/common/management/commands/import_districts.py
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from apps.common.models import Region, District

class Command(BaseCommand):
    help = "Импорт районов (District) из XLSX. Ожидаемые колонки: region_code, region_name, district_number, district_name"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Путь к XLSX-файлу")

    def handle(self, *args, **opts):
        path = opts["xlsx_path"]
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        # ожидаем имена колонок
        try:
            i_code = header.index("region_code")
            i_dnum = header.index("district_number")
            i_dname = header.index("district_name")
        except ValueError:
            raise CommandError("В первой строке должны быть колонки: region_code, district_number, district_name")

        created, updated, skipped = 0, 0, 0
        for row in ws.iter_rows(min_row=2):
            reg_code = (row[i_code].value or "").strip().upper()
            dnum = row[i_dnum].value
            dname = (row[i_dname].value or "").strip()
            if not (reg_code and dnum and dname):
                skipped += 1
                continue
            try:
                region = Region.objects.get(code=reg_code)
            except Region.DoesNotExist:
                self.stderr.write(self.style.WARNING(f"Регион с кодом {reg_code} не найден — пропуск"))
                skipped += 1
                continue
            obj, is_created = District.objects.update_or_create(
                region=region, number=int(dnum),
                defaults={"name": dname}
            )
            created += int(is_created)
            updated += int(not is_created)

        self.stdout.write(self.style.SUCCESS(f"Готово: создано {created}, обновлено {updated}, пропущено {skipped}"))
